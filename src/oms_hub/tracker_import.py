import hashlib
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from oms_hub.repositories import CatalogRepository, LectureInput

_SECTION = re.compile(r"^(?P<subject>.+?)\s+(?P<exam>\d+)\s*$")
_SKIP_SHEETS = {
    "Instructions!",
    "GRADES",
    "EXAM SHEET EXAMPLE",
    "EXAM DATES",
}
_ALIASES = {
    "HEME": "Heme/Lymph",
    "HEME / LYMPH": "Heme/Lymph",
    "HEME/LYMPH": "Heme/Lymph",
    "NEURO": "Neuro",
    "MSK": "MSK",
    "CARDIO": "Cardio",
    "RENAL": "Renal",
    "RESP": "Resp",
    "OPP": "OPP",
    "EPC": "EPC",
}


@dataclass(frozen=True, slots=True)
class TrackerImportResult:
    imported: int
    updated: int
    issues: int
    source_sha256: str


@dataclass(frozen=True, slots=True)
class TrackerParseIssue:
    sheet: str
    row_number: int
    message: str
    raw_values: str


@dataclass(frozen=True, slots=True)
class ParsedTracker:
    lectures: tuple[LectureInput, ...]
    issues: tuple[TrackerParseIssue, ...]
    source_sha256: str
    source_name: str


def _normalized_subject(value: str) -> str:
    key = re.sub(r"\s+", " ", value.strip().upper())
    key = re.sub(r"\s*-\s*EXAM$", "", key)
    return _ALIASES.get(key, value.strip().title())


def _exam_dates(sheet: Any) -> dict[tuple[str, int], str]:
    result: dict[tuple[str, int], str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_date, raw_exam = row[0], row[1]
        if not raw_date or not raw_exam:
            continue
        value = raw_date.date() if isinstance(raw_date, datetime) else raw_date
        if not isinstance(value, date):
            continue
        for part in str(raw_exam).split("/"):
            match = _SECTION.match(part.strip())
            if match:
                subject = _normalized_subject(match.group("subject"))
                result[(subject, int(match.group("exam")))] = value.isoformat()
    return result


def parse_tracker(path: Path) -> ParsedTracker:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        exam_dates = (
            _exam_dates(workbook["EXAM DATES"])
            if "EXAM DATES" in workbook.sheetnames
            else {}
        )
        issues: list[TrackerParseIssue] = []
        lectures: list[LectureInput] = []
        seen: dict[tuple[str, int, int], tuple[str, int, str]] = {}
        reported_first_duplicates: set[tuple[str, int, int]] = set()

        for sheet in workbook.worksheets:
            if sheet.title in _SKIP_SHEETS:
                continue
            subject: str | None = None
            exam_number: int | None = None
            for row_number, row in enumerate(
                sheet.iter_rows(values_only=True),
                start=1,
            ):
                first = str(row[0]).strip() if row and row[0] is not None else ""
                second = (
                    str(row[1]).strip()
                    if len(row) > 1 and row[1] is not None
                    else ""
                )
                third = (
                    str(row[2]).strip()
                    if len(row) > 2 and row[2] is not None
                    else ""
                )
                section = _SECTION.match(first) if not second else None
                if section:
                    subject = _normalized_subject(section.group("subject"))
                    exam_number = int(section.group("exam"))
                    continue
                if (
                    not first
                    or first == "#"
                    or subject is None
                    or exam_number is None
                ):
                    continue
                if not first.isdigit():
                    issues.append(
                        TrackerParseIssue(
                            sheet.title,
                            row_number,
                            "ambiguous lecture number",
                            json.dumps(row[:3], default=str),
                        )
                    )
                    continue
                if not second:
                    issues.append(
                        TrackerParseIssue(
                            sheet.title,
                            row_number,
                            "missing lecture title",
                            json.dumps(row[:3], default=str),
                        )
                    )
                    continue
                key = (subject, exam_number, int(first))
                if key in seen:
                    first_sheet, first_row, first_raw = seen[key]
                    if key not in reported_first_duplicates:
                        issues.append(
                            TrackerParseIssue(
                                first_sheet,
                                first_row,
                                f"duplicate lecture number also appears at "
                                f"{sheet.title} row {row_number}",
                                first_raw,
                            )
                        )
                        reported_first_duplicates.add(key)
                    issues.append(
                        TrackerParseIssue(
                            sheet.title,
                            row_number,
                            f"duplicate lecture number first appears at "
                            f"{first_sheet} row {first_row}",
                            json.dumps(row[:3], default=str),
                        )
                    )
                    continue
                seen[key] = (
                    sheet.title,
                    row_number,
                    json.dumps(row[:3], default=str),
                )
                lectures.append(
                    LectureInput(
                        subject=subject,
                        exam_number=exam_number,
                        lecture_number=int(first),
                        topic=second,
                        lecturer=third,
                        exam_date=exam_dates.get((subject, exam_number)),
                    )
                )

        return ParsedTracker(
            lectures=tuple(lectures),
            issues=tuple(issues),
            source_sha256=digest,
            source_name=path.name,
        )
    finally:
        workbook.close()


class TrackerImporter:
    def __init__(self, repository: CatalogRepository):
        self.repository = repository

    def import_once(self, path: Path) -> TrackerImportResult:
        parsed = parse_tracker(path)
        if self.repository.has_import_hash(parsed.source_sha256):
            raise ValueError("tracker workbook has already been imported")

        imported, updated = self.repository.commit_tracker_import(
            list(parsed.lectures),
            [
                (
                    issue.sheet,
                    issue.row_number,
                    issue.message,
                    issue.raw_values,
                )
                for issue in parsed.issues
            ],
            parsed.source_sha256,
            parsed.source_name,
        )
        return TrackerImportResult(
            imported=imported,
            updated=updated,
            issues=len(parsed.issues),
            source_sha256=parsed.source_sha256,
        )

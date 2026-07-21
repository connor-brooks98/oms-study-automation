from datetime import date

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from oms_hub.models import LectureModel
from oms_hub.repositories import CatalogRepository
from oms_hub.tracker_import import TrackerImporter


def test_imports_combined_exam_sheet_and_records_ambiguous_number(database, tmp_path):
    path = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    dates = workbook.active
    dates.title = "EXAM DATES"
    dates.append(["DATE:", "BLOCK EXAM:"])
    dates.append([date(2026, 7, 3), "NEURO 1 / MSK 1 / HEME 1"])
    sheet = workbook.create_sheet("NEURO 1  MSK 1  HEME 1")
    sheet.append(["NEURO 1"])
    sheet.append(["#", "Lecture Title", "Lecturer"])
    sheet.append([1, "General CNS Pathology", "Teresa Campbell, MD"])
    sheet.append(["19?", "Sleep Disorders", "Leah Snodgrass, MD"])
    sheet.append(["HEME / LYMPH 1"])
    sheet.append(["#", "Lecture Title", "Lecturer"])
    sheet.append([4, "Anemia I", "Jun Wang, MD, PhD"])
    workbook.save(path)

    repo = CatalogRepository(database)
    result = TrackerImporter(repo).import_once(path)
    lectures = repo.list_lectures()

    assert result.imported == 2
    assert result.issues == 1
    assert [(item.subject, item.exam_number, item.lecture_number) for item in lectures] == [
        ("Heme/Lymph", 1, 4),
        ("Neuro", 1, 1),
    ]
    assert {item.exam_date for item in lectures} == {"2026-07-03"}


def test_second_import_with_same_hash_is_rejected(database, tmp_path):
    path = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "EXAM DATES"
    workbook.save(path)
    importer = TrackerImporter(CatalogRepository(database))
    importer.import_once(path)

    with pytest.raises(ValueError, match="tracker workbook has already been imported"):
        importer.import_once(path)


def test_normalizes_section_headers_that_include_exam_label(database, tmp_path):
    path = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OPP 1  EPC 1"
    sheet.append(["OPP - Exam 1"])
    sheet.append(["#", "Lecture Title", "Lecturer"])
    sheet.append([1, "Course Introduction", "Dr. Example"])
    workbook.save(path)

    repository = CatalogRepository(database)
    TrackerImporter(repository).import_once(path)

    assert repository.list_lectures()[0].subject == "OPP"


def test_database_transaction_rolls_back_partial_catalog_write(database):
    with pytest.raises(RuntimeError, match="forced rollback"):
        with database.session() as session:
            session.add(
                LectureModel(
                    subject="Neuro",
                    exam_number=1,
                    lecture_number=1,
                    topic="General CNS Pathology",
                    lecturer="Teresa Campbell, MD",
                    exam_date="2026-07-03",
                )
            )
            session.flush()
            raise RuntimeError("forced rollback")

    with database.session() as session:
        count = session.scalar(
            select(func.count()).select_from(LectureModel)
        )
    assert count == 0
